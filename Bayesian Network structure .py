from pylab import *
import matplotlib.pyplot as plt
from IPython.display import display, HTML
import os
import pyAgrum as gum
import math
import pyAgrum.lib.explain as explain

from azure.core.credentials import AzureKeyCredential
from azure.ai.language.conversations import ConversationAnalysisClient

from azure.ai.textanalytics import TextAnalyticsClient
from azure.core.credentials import AzureKeyCredential

import openpyxl


#--------------------------------------------------------------------------------------------------------------------------------------------------------------------------------

bn=gum.BayesNet('PsyMonitor')
g = bn.add(gum.LabelizedVariable('Gender','Con quale genere ti identifichi?',2)) # Man, woman
#bn.cpt('Gender').fillWith(gum.randomDistribution(2))
d = bn.add(gum.LabelizedVariable('Department','In quale dipartimento studi?',26)) # https://www.international.unina.it/departments/

c = bn.add(gum.LabelizedVariable('Course', 'Quale corso frequenti?', 173)) # In the Federico II university, there are 73 bachelor courses, 66 master course and 8 single cycle courses. Let's take the DIETI as an example (https://www.dieti.unina.it/index.php/it/studenti/orientamento-2)
bn.addArc(c,g)
bn.addArc(c,d)



#bn.cpt('Department')
#bn.cpt('Course')
#bn.cpt('Course')

#

dataframe = openpyxl.load_workbook('C:/Users/waari/Desktop/Iscritti in Corsi attivi 22-23.xlsx')
dataframe1 = dataframe.active

male_list = []
# Females
for row in range(4, dataframe1.max_row-1):
    for col in dataframe1.iter_cols(6, dataframe1.max_column-2):
        male_list.append(col[row].value)
        # print(col[row].value)
print('male_list: ')
print(male_list)

female_list = []
# Females
for row in range(4, dataframe1.max_row-1):
    for col in dataframe1.iter_cols(7, dataframe1.max_column-1):
        female_list.append(col[row].value)
        # print(col[row].value)
print('female_list: ')
print(female_list)

c=0
for (i,j) in zip(male_list,female_list):
    bn.cpt("Gender")[c, :] = [i, j]
    c+=1
    if ( c == len(male_list) ):
        break

bn.cpt('Gender').normalizeAsCPT()

gen_list = []
start = True
for row in range(4, dataframe1.max_row - 1):
    print('')
    list = []
    for col in dataframe1.iter_cols(2, dataframe1.max_column - 3):
        print(col[row].value)
        list.append(col[row].value)

    del list[1:2]
    del list[1:2]

    if list[0] == None:
        courses.append(list[1])
    else:
        if not start:
            gen_list.append((dept, courses))
        else:
            start = False
        dept = list[0]
        courses = []
        courses.append(list[1])
gen_list.append((dept, courses))

for elem in gen_list:
    dip = elem[0]
    lista_corsi_perdip = elem[1]
    for corso in lista_corsi_perdip:
        bn.cpt("Department")[{'Department': elem[0], 'Course': corso}] = 1

bn.cpt('Department')
bn.cpt('Department').normalizeAsCPT()

pass

## next part of the code is not tested yet


ie=gum.LazyPropagation(bn)
ie.makeInference()

print('Entropy of g is', ie.posterior('Gender').entropy())
print('Entropy of d is', ie.posterior('Department').entropy())
print('Entropy of c is', ie.posterior('Course').entropy())
ie.makeInference()

print(bn.variableFromName('Course').description())
answer = "Sono iscritto a Informatica"

course = zeros(147)
course[0] = 1
ie.setEvidence({'Course': course})
print('Entropy of g is', ie.posterior('Gender').entropy())
print('Entropy of d is', ie.posterior('Department').entropy())
print('Entropy of c is', ie.posterior('Course').entropy())
gnb.showInference(bn,evs={'Course': course})

print(bn.variableFromName('Department').description())
dept = zeros(26)
dept[9] = 1
ie.setEvidence({'Department': dept})
ie.eraseEvidence('Course')
refused = []
refused.append('Course')
print('Entropy of g is', ie.posterior('Gender').entropy())
print('Entropy of d is', ie.posterior('Department').entropy())
print('Entropy of c is', ie.posterior('Course').entropy())
gnb.showInference(bn,evs={'Department': dept})

p = bn.add(gum.LabelizedVariable('p', 'Personality', 5))

bn.cpt('p').fillWith(1)
bn.cpt('p').normalizeAsCPT()

var = gum.DiscretizedVariable('s1', 'Subtopic1')
var.addTick(0)
var.addTick(0.33)
var.addTick(0.66)
var.addTick(1)

s1 = bn.add(var)

var = gum.DiscretizedVariable('s2', 'Subtopic1')
var.addTick(0)
var.addTick(0.33)
var.addTick(0.66)
var.addTick(1)

s2 = bn.add(var)

var = gum.DiscretizedVariable('s3', 'Subtopic1')
var.addTick(0)
var.addTick(1/3)
var.addTick(2/3)
var.addTick(1)

s3 = bn.add(var)

var = gum.DiscretizedVariable('s4', 'Subtopic1')
var.addTick(0)
var.addTick(1/3)
var.addTick(2/3)
var.addTick(1)

s4 = bn.add(var)

var = gum.DiscretizedVariable('t1', 'Topic1')
var.addTick(0)
var.addTick(1/3)
var.addTick(2/3)
var.addTick(1)

t1 = bn.addMEDIAN(var)

var = gum.DiscretizedVariable('t2', 'Topic2')
var.addTick(0)
var.addTick(1/3)
var.addTick(2/3)
var.addTick(1)

t2 = bn.addMEDIAN(var)

var = gum.DiscretizedVariable('gf', 'General Feedback')
var.addTick(0)
var.addTick(1/3)
var.addTick(2/3)
var.addTick(1)
gf = bn.addMEDIAN(var)

bn.addArc(s1,t1)
bn.addArc(s2,t1)
bn.addArc(s3,t2)
bn.addArc(s4,t2)
bn.addArc(t1,gf)
bn.addArc(t2,gf)

bn.addArc(c,s1)
bn.addArc(c,s2)
bn.addArc(c,s3)
bn.addArc(c,s4)
bn.addArc(p,s1)
bn.addArc(p,s2)
bn.addArc(p,s3)
bn.addArc(p,s4)

bn.cpt('s1').fillWith(1).normalizeAsCPT()
bn.cpt('s2').fillWith(1).normalizeAsCPT()
bn.cpt('s3').fillWith(1).normalizeAsCPT()
bn.cpt('s4').fillWith(1).normalizeAsCPT()

ie.eraseEvidence('Department')
evidences= {}
print('Entropy of g is', ie.posterior('Gender').entropy())
print('Entropy of d is', ie.posterior('Department').entropy())
print('Entropy of c is', ie.posterior('Course').entropy())

ie.setEvidence({'Course': course})
print('Entropy of g is', ie.posterior('Gender').entropy())
print('Entropy of d is', ie.posterior('Department').entropy())
print('Entropy of c is', ie.posterior('Course').entropy())


clu_endpoint = "https://psymonitor.cognitiveservices.azure.com/"
clu_key = "fb821673335b408f9b324740f5c3291a"
project_name = "PsyMonitor"
deployment_name = "PsyMonitor"

client = ConversationAnalysisClient(clu_endpoint, AzureKeyCredential(clu_key))

def personalityFeedbackAnalyse(query):
  with client:
    return client.analyze_conversation(
        task={
            "kind": "Conversation",
            "analysisInput": {
                "conversationItem": {
                    "participantId": "1",
                    "id": "1",
                    "modality": "text",
                    "language": "it",
                    "text": query
                },
                "isLoggingEnabled": False
            },
            "parameters": {
                "projectName": project_name,
                "deploymentName": deployment_name,
                "verbose": True
            }
        }
    )

result = personalityFeedbackAnalyse('Direi di si')
print("query: {}".format(result["result"]["query"]))
print("top intent: {}".format(result["result"]["prediction"]["topIntent"]))
for intent in result["result"]["prediction"]["intents"]:
  print("{}: {}".format(intent['category'], intent['confidenceScore']))

#---------------------------------------------------------

language_key = clu_key
language_endpoint = clu_endpoint

def analyseSentiment(text):
  ta_credential = AzureKeyCredential(language_key)
  text_analytics_client = TextAnalyticsClient(
          endpoint=language_endpoint,
          credential=ta_credential)
  return text_analytics_client.analyze_sentiment([text], show_opinion_mining=True)[0]


print()
result = analyseSentiment("Non mi piace il corso di programmazione. Il prof mi odia.")
print("Negative: {}".format(result.confidence_scores.negative))
print("Neutral: {}".format(result.confidence_scores.neutral))
print("Positive: {}".format(result.confidence_scores.positive))
for mined_opinion in result.sentences[0].mined_opinions:
  target = mined_opinion.target
  print("'{}' target '{}'".format(target.sentiment, target.text))
  print("Target score:\nPositive={0:.2f}\nNegative={1:.2f}\n".format(
      target.confidence_scores.positive,
      target.confidence_scores.negative,
  ))
  for assessment in mined_opinion.assessments:
      print("'{}' assessment '{}'".format(assessment.sentiment, assessment.text))
      print("Assessment score:\nPositive={0:.2f}\nNegative={1:.2f}\n".format(
          assessment.confidence_scores.positive,
          assessment.confidence_scores.negative,
      ))
  print("\n")